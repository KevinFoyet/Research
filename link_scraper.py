import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def get_coin_link(coin_name):
    if coin_name is None:
        print(f"No coin name found.")
        return "No coin name"
    url = f"https://icomarks.com/ico/{coin_name.replace(' ', '-').lower()}"
    response = requests.get(url)
    if response.status_code == 200:
        print(f"Link found for {coin_name}: {url}")
        return url
    print(f"No link found for {coin_name}.")
    return "No link found"


def main():
    # Load the workbook
    workbook = load_workbook('testing_Descrip.xlsx')

    # Select the main sheet
    sheet = workbook.active

    # Find the index of the coin name column (assume it's the first column)
    name_column_index = 1  # 1-based index

    # Iterate over the coin names and add the links
    for row in range(2, sheet.max_row + 1):  # start from 2 to skip the header
        coin_name = sheet.cell(row=row, column=name_column_index).value
        link = get_coin_link(coin_name)
        
        # Assuming the link should be added to the next column
        sheet.cell(row=row, column=name_column_index + 1).value = link

    # Save the workbook
    workbook.save('testing_Descrip.xlsx')

if __name__ == "__main__":
    main()
