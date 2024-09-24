import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import urllib.parse
import time

def process_excel_file(file_path):
    df = pd.read_excel(file_path)
    wb = load_workbook(filename=file_path)
    ws = wb.active

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    name_col_idx = 2 - 1
    link_col_idx = 51 - 1
    whitepaper_col_idx = 49 - 1

    base_url = 'https://www.allcryptowhitepapers.com/'
    
    rows_per_batch = 50  # number of rows to process before pausing
    pause_duration = 60  # pause duration in seconds

    for index, row in enumerate(df.itertuples(), start=1):
        company_name = row[name_col_idx+1]
        try:
            search_url = urllib.parse.urljoin(base_url, company_name.lower().replace(' ', '-') + '-whitepaper/')
            response = requests.get(search_url)
            if response.status_code == 200:
                df.at[row.Index, df.columns[link_col_idx]] = search_url
                df.at[row.Index, df.columns[whitepaper_col_idx]] = 'Found'
                ws.cell(row=index+1, column=link_col_idx+1, value=search_url)
                ws.cell(row=index+1, column=whitepaper_col_idx+1, value='Found').fill = green_fill
            else:
                df.at[row.Index, df.columns[link_col_idx]] = ''
                df.at[row.Index, df.columns[whitepaper_col_idx]] = 'Not Found'
                ws.cell(row=index+1, column=link_col_idx+1, value='')
                ws.cell(row=index+1, column=whitepaper_col_idx+1, value='Not Found').fill = red_fill
        except Exception as e:
            print(f"Error processing company '{company_name}': {str(e)}")

        if index % rows_per_batch == 0:
            time.sleep(pause_duration)

    df.to_excel(file_path, index=False)
    wb.save(file_path)



# usage
process_excel_file(r'C:\Users\Kevin\OneDrive - Cal Poly Pomona\Documents\combined_2017-2018_Updated_05092023.xlsx')


